#!/usr/bin/env python3
"""
grader.py - Git-based authenticity triage for student GitHub Classroom repos
"""

import argparse
import csv
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path
from statistics import mean, stdev

try:
    import tomllib
except ImportError:
    sys.exit("Python 3.11+ required for tomllib")


# ---------------------------------------------------------------------------
# Git helpers
# ---------------------------------------------------------------------------

def git(repo_path, *args):
    result = subprocess.run(
        ["git", "-C", str(repo_path)] + list(args),
        capture_output=True, text=True
    )
    return result.stdout.strip()


# Exclude bot commits (github-classroom[bot], github-actions[bot], etc.)
_STUDENT_FILTER = ("--perl-regexp", r"--author=^((?!\[bot\]).)*$")

def git_log(repo_path, *args):
    """git log restricted to student commits — bots excluded."""
    return git(repo_path, "log", *_STUDENT_FILTER, *args)


def commit_dates(repo_path):
    out = git_log(repo_path, "--format=%ct")
    if not out:
        return []
    return [datetime.fromtimestamp(int(ts), tz=timezone.utc) for ts in out.splitlines()]


def spread_days(dates):
    return len(set(d.date() for d in dates))


def interval_cv(dates):
    """Coefficient of variation of inter-commit intervals. High = irregular (natural)."""
    if len(dates) < 3:
        return 0.0
    intervals = [abs((dates[i] - dates[i+1]).total_seconds()) for i in range(len(dates)-1)]
    intervals = [x for x in intervals if x > 0]
    if len(intervals) < 2 or mean(intervals) == 0:
        return 0.0
    return stdev(intervals) / mean(intervals)


def deletion_ratio(repo_path):
    out = git_log(repo_path, "--shortstat", "--format=")
    total_ins = total_del = 0
    for line in out.splitlines():
        if m := re.search(r'(\d+) insertion', line):
            total_ins += int(m.group(1))
        if m := re.search(r'(\d+) deletion', line):
            total_del += int(m.group(1))
    return (total_del / total_ins) if total_ins else 0.0


def max_commit_insertion_ratio(repo_path):
    """Fraction of total insertions that landed in the single largest commit."""
    out = git_log(repo_path, "--shortstat", "--format=%H")
    commit_ins = []
    current = 0
    for line in out.splitlines():
        line = line.strip()
        if not line:
            commit_ins.append(current)
            current = 0
        elif m := re.search(r'(\d+) insertion', line):
            current = int(m.group(1))
    if current:
        commit_ins.append(current)
    total = sum(commit_ins)
    return (max(commit_ins) / total) if total and commit_ins else 0.0


def session_file_churn(repo_path, session_gap_hours=4):
    """Count files modified in 3+ distinct sessions."""
    out = git_log(repo_path, "--format=%ct", "--name-only")
    file_times = {}
    current_ts = None
    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        if line.isdigit():
            current_ts = int(line)
        elif current_ts is not None:
            file_times.setdefault(line, []).append(current_ts)

    count = 0
    for timestamps in file_times.values():
        if len(timestamps) < 3:
            continue
        timestamps.sort()
        sessions = 1
        for i in range(1, len(timestamps)):
            if (timestamps[i] - timestamps[i-1]) > session_gap_hours * 3600:
                sessions += 1
        if sessions >= 3:
            count += 1
    return count


def lazy_message_count(repo_path):
    out = git_log(repo_path, "--format=%s")
    pattern = re.compile(
        r'^(wip|fix|update|changes|stuff|asdf|test|commit|done|final|more|misc|lol|idk|a+|\.+)$',
        re.I
    )
    return sum(1 for line in out.splitlines() if pattern.match(line.strip()))


def total_window_hours(dates):
    if len(dates) < 2:
        return 0
    return abs((dates[0] - dates[-1]).total_seconds()) / 3600


def started_early(dates, assigned_date, due_date):
    """True if any commit happened before the last 20% of the assignment window."""
    if not dates:
        return False
    window = (due_date - assigned_date).total_seconds()
    threshold = assigned_date.timestamp() + window * 0.80
    return any(d.timestamp() < threshold for d in dates)


def cleanup_commit_count(repo_path, window_hours=24, deletion_ratio_threshold=3.0):
    """Count commits that look like LLM cleanup: deletion-heavy commit occurring
    within window_hours after a large insertion commit.

    Pattern: student pastes generated code (big insertion), then quickly removes
    comments, dead code, or formatting artifacts (big deletion, few insertions).
    """
    out = git_log(repo_path, "--shortstat", "--format=%ct")
    if not out:
        return 0

    # Build list of (timestamp, insertions, deletions) per commit
    commits = []
    current_ts = None
    for line in out.splitlines():
        line = line.strip()
        if not line:
            continue
        if line.isdigit():
            current_ts = int(line)
        else:
            ins = int(m.group(1)) if (m := re.search(r'(\d+) insertion', line)) else 0
            dels = int(m.group(1)) if (m := re.search(r'(\d+) deletion', line)) else 0
            if current_ts is not None:
                commits.append((current_ts, ins, dels))
                current_ts = None

    count = 0
    for i, (ts, ins, dels) in enumerate(commits):
        # Deletion-heavy commit (far more deletions than insertions)
        if ins == 0 or dels / max(ins, 1) < deletion_ratio_threshold:
            continue
        # Check if a large insertion commit preceded it within the window
        for j in range(i + 1, len(commits)):
            prev_ts, prev_ins, _ = commits[j]
            if (ts - prev_ts) > window_hours * 3600:
                break
            if prev_ins > 50:  # meaningful insertion threshold
                count += 1
                break

    return count


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

def score_repo(repo_path, config):
    assigned = datetime.fromisoformat(config["assignment"]["assigned_date"]).replace(tzinfo=timezone.utc)
    due      = datetime.fromisoformat(config["assignment"]["due_date"]).replace(tzinfo=timezone.utc)
    expected = config["assignment"].get("expected_commits", 5)
    thresholds_cfg = config.get("thresholds", {})
    t = {
        "min_spread_days":  thresholds_cfg.get("min_spread_days",  3),
        "window_cap_hours": thresholds_cfg.get("window_cap_hours", 2),
    }
    w = {
        "commit_count":     15,
        "spread_days":      20,
        "started_early":    15,
        "interval_cv":      10,
        "deletions":        15,
        "file_churn":       10,
        "no_dump":          15,
        "cleanup_commits":   0,  # off by default; enable in [weights] for term projects
        **config.get("weights", {}),
    }

    dates = commit_dates(repo_path)
    count = len(dates)

    if count == 0:
        empty_stats = {"commits": 0, "spread": 0, "early": False, "cv": 0.0,
                       "del_pct": 0.0, "churn": 0, "dump_pct": 0.0, "cleanups": 0}
        return 0, "MISSING: no commits found", empty_stats

    found, missing = [], []
    points = 0

    # commit count reasonable
    if count >= max(expected * 0.5, 2):
        points += w["commit_count"]
        found.append(f"{count} commits (expected ~{expected})")
    else:
        missing.append(f"only {count} commits (expected ~{expected})")

    # spread across min_spread_days+
    days = spread_days(dates)
    min_days = t["min_spread_days"]
    if days >= min_days:
        points += w["spread_days"]
        found.append(f"work spread across {days} days")
    else:
        missing.append(f"all work within {days} day(s) (need {min_days})")

    # started before final 20% of window
    if started_early(dates, assigned, due):
        points += w["started_early"]
        found.append("started before final stretch of deadline")
    else:
        missing.append("no commits before final 20% of assignment window")

    # irregular commit intervals (natural rhythm)
    cv = interval_cv(dates)
    if cv > 0.5:
        points += w["interval_cv"]
        found.append(f"irregular commit intervals (CV={cv:.2f})")
    else:
        missing.append(f"uniform commit intervals (CV={cv:.2f})")

    # deletions present (code was revised)
    dratio = deletion_ratio(repo_path)
    if dratio > 0.05:
        points += w["deletions"]
        found.append(f"deletions present ({dratio:.0%} of insertions)")
    else:
        missing.append(f"few/no deletions ({dratio:.0%} of insertions)")

    # file churn across sessions
    churn = session_file_churn(repo_path)
    if churn >= 1:
        points += w["file_churn"]
        found.append(f"{churn} file(s) revised across multiple sessions")
    else:
        missing.append("no files revised across multiple sessions")

    # no single-commit dump
    dump = max_commit_insertion_ratio(repo_path)
    if dump < 0.70:
        points += w["no_dump"]
        found.append(f"no single-commit dump (largest commit: {dump:.0%} of insertions)")
    else:
        missing.append(f"single commit contains {dump:.0%} of all insertions")

    # cleanup commits (deletion-heavy commit after large insertion = paste cleanup)
    cleanups = cleanup_commit_count(repo_path)
    if w["cleanup_commits"] > 0:
        if cleanups == 0:
            points += w["cleanup_commits"]
            found.append("no cleanup-after-paste commits detected")
        else:
            missing.append(f"{cleanups} cleanup commit(s) detected (deletion-heavy commit after large insertion)")

    # Cap: all commits within window_cap_hours is a hard red flag (0 = disabled)
    cap_hours = t["window_cap_hours"]
    if cap_hours > 0:
        hours = total_window_hours(dates)
        if hours < cap_hours and count > 2:
            points = min(points, 20)
            missing.append(f"all {count} commits within {hours:.1f} hours (score capped at 20)")

    parts = []
    if found:
        parts.append("FOUND: " + "; ".join(found))
    if missing:
        parts.append("MISSING: " + "; ".join(missing))

    stats = {
        "commits":    count,
        "spread":     days,
        "early":      started_early(dates, assigned, due),
        "cv":         cv,
        "del_pct":    dratio,
        "churn":      churn,
        "dump_pct":   dump,
        "cleanups":   cleanups,
    }

    return points, " | ".join(parts), stats


def triage(score, pass_t, flag_t):
    if score >= pass_t:
        return "PASS"
    if score <= flag_t:
        return "FLAG"
    return "REVIEW"


# ---------------------------------------------------------------------------
# I/O helpers
# ---------------------------------------------------------------------------

def repo_name_from_url(url):
    """Works for both HTTPS and SSH git URLs."""
    name = url.rstrip("/").split("/")[-1]
    return name.removesuffix(".git")


def find_local_path(repo_name, repos_dir):
    if repos_dir:
        p = Path(repos_dir) / repo_name
        if p.exists():
            return p
    for subdir in Path(".").glob("*-submissions"):
        p = subdir / repo_name
        if p.exists():
            return p
    return None


def load_roster(path):
    if not path or not Path(path).exists():
        return {}
    with open(path) as f:
        return {row["github_username"]: row["display_name"] for row in csv.DictReader(f)}


# ---------------------------------------------------------------------------
# Spreadsheet output
# ---------------------------------------------------------------------------

def write_md(rows, path, config):
    ICONS = {"FLAG": "🚩", "REVIEW": "🔍", "PASS": "✅"}
    counts = {k: sum(1 for r in rows if r["triage"] == k) for k in ["FLAG", "REVIEW", "PASS"]}
    name = config["assignment"].get("name", "Results")

    lines = [
        f"# {name}",
        "",
        f"| | Count |",
        f"|---|---|",
        f"| {ICONS['FLAG']} FLAG   | {counts['FLAG']} |",
        f"| {ICONS['REVIEW']} REVIEW | {counts['REVIEW']} |",
        f"| {ICONS['PASS']} PASS   | {counts['PASS']} |",
        "",
    ]

    for bucket in ["FLAG", "REVIEW", "PASS"]:
        bucket_rows = [r for r in rows if r["triage"] == bucket]
        if not bucket_rows:
            continue
        lines.append(f"## {ICONS[bucket]} {bucket} ({len(bucket_rows)})")
        lines.append("")
        lines.append("| Name | Score | Grade | Reasoning |")
        lines.append("|------|------:|-------|-----------|")
        for r in bucket_rows:
            reasoning = r["reasoning"].replace("|", "\\|")
            lines.append(f"| [{r['name']}]({r['repo_url']}) | {r['score']} | {r['grade']} | {reasoning} |")
        lines.append("")

    path.write_text("\n".join(lines))


def write_xlsx(rows, path, config):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    COLORS = {
        "FLAG":   "FFCCCC",  # red tint
        "REVIEW": "FFF2CC",  # yellow tint
        "PASS":   "CCFFCC",  # green tint
    }
    HEADER_FILL  = "2F4F7F"
    STAT_FILL    = "3A6599"   # slightly lighter blue for stat header group

    # Core columns + stat columns + reasoning
    CORE_COLS    = ["name", "repo_url", "triage", "score", "grade"]
    CORE_HEADERS = ["Name", "Repository", "Triage", "Score", "Grade"]
    STAT_KEYS    = ["commits", "spread", "early", "cv", "del_pct", "churn", "dump_pct", "cleanups"]
    STAT_HEADERS = ["Commits", "Spread\n(days)", "Early", "CV", "Del%", "Churn\n(files)", "Dump%", "Cleanups"]
    TRAIL_COLS    = ["reasoning"]
    TRAIL_HEADERS = ["Reasoning"]

    ALL_COLS    = CORE_COLS + STAT_KEYS + TRAIL_COLS
    ALL_HEADERS = CORE_HEADERS + STAT_HEADERS + TRAIL_HEADERS

    wb = Workbook()
    ws = wb.active
    ws.title = config["assignment"].get("name", "Results")

    # Header row
    thin = Side(style="thin", color="AAAAAA")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col_i, (key, header) in enumerate(zip(ALL_COLS, ALL_HEADERS), 1):
        is_stat = key in STAT_KEYS
        cell = ws.cell(row=1, column=col_i, value=header)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor=STAT_FILL if is_stat else HEADER_FILL)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"

    # Data rows
    for row_i, row in enumerate(rows, 2):
        fill  = PatternFill("solid", fgColor=COLORS[row["triage"]])
        stats = row.get("stats", {})

        def stat_display(key):
            v = stats.get(key)
            if v is None:
                return ""
            if key == "early":
                return "Y" if v else "N"
            if key in ("del_pct", "dump_pct"):
                return f"{v:.0%}"
            if key == "cv":
                return f"{v:.2f}"
            return v  # int for commits, spread, churn, cleanups

        col_vals = (
            [row[k] for k in CORE_COLS]
            + [stat_display(k) for k in STAT_KEYS]
            + [row["reasoning"]]
        )

        for col_i, (key, val) in enumerate(zip(ALL_COLS, col_vals), 1):
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.fill   = fill
            cell.border = border
            cell.alignment = Alignment(
                vertical="top",
                wrap_text=(key == "reasoning"),
                horizontal="center" if key in STAT_KEYS + ["score", "triage"] else "left",
            )
            if key == "repo_url":
                cell.hyperlink = val if val.startswith("http") else f"https://github.com/{val.split(':')[1].removesuffix('.git')}"
                cell.font = Font(color="0563C1", underline="single")

    # Column widths
    widths = {
        "name": 24, "repo_url": 36, "triage": 9, "score": 7, "grade": 8,
        "commits": 9, "spread": 8, "early": 7, "cv": 6,
        "del_pct": 7, "churn": 8, "dump_pct": 7, "cleanups": 9,
        "reasoning": 70,
    }
    for col_i, key in enumerate(ALL_COLS, 1):
        ws.column_dimensions[get_column_letter(col_i)].width = widths.get(key, 12)

    # Row heights
    for row_i in range(2, len(rows) + 2):
        ws.row_dimensions[row_i].height = 15

    # Summary sheet with pie chart
    counts = {"FLAG": 0, "REVIEW": 0, "PASS": 0}
    for row in rows:
        counts[row["triage"]] += 1

    from openpyxl.chart import PieChart, Reference
    from openpyxl.chart.series import DataPoint

    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Triage"
    ws2["B1"] = "Count"
    for i, (label, count) in enumerate(counts.items(), 2):
        ws2.cell(row=i, column=1, value=label)
        ws2.cell(row=i, column=2, value=count)

    chart = PieChart()
    chart.title = config["assignment"].get("name", "Results")
    chart.style = 10

    data   = Reference(ws2, min_col=2, min_row=1, max_row=4)
    labels = Reference(ws2, min_col=1, min_row=2, max_row=4)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(labels)
    chart.dataLabels = chart.dataLabels or __import__("openpyxl.chart.label", fromlist=["DataLabelList"]).DataLabelList()
    chart.dataLabels.showPercent = True
    chart.dataLabels.showCatName = True
    chart.dataLabels.showVal = False

    # Match slice colors to row colors: FLAG=red, REVIEW=yellow, PASS=green
    from openpyxl.drawing.fill import PatternFillProperties
    slice_colors = ["FF4444", "FFCC00", "44BB44"]
    for idx, color in enumerate(slice_colors):
        pt = DataPoint(idx=idx)
        pt.graphicalProperties.solidFill = color
        chart.series[0].dPt.append(pt)

    chart.width  = 15
    chart.height = 12
    ws2.add_chart(chart, "D1")

    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(description="Git authenticity grader for student repos")
    parser.add_argument("--config",    required=True, help="project.toml")
    parser.add_argument("--repos",     required=True, help="repos.txt (one URL per line)")
    parser.add_argument("--roster",    help="roster.csv (github_username,display_name)")
    parser.add_argument("--repos-dir", help="Directory containing cloned repos")
    parser.add_argument("--skip-clone", action="store_true", help="Skip cloning; use existing local repos")
    parser.add_argument("--output",    default="results.csv", help="Output CSV (default: results.csv)")
    args = parser.parse_args()

    with open(args.config, "rb") as f:
        config = tomllib.load(f)

    pass_t = config["thresholds"]["pass"]
    flag_t = config["thresholds"]["flag"]

    with open(args.repos) as f:
        urls = [line.strip() for line in f if line.strip()]

    roster = load_roster(args.roster)
    prefix = config["assignment"].get("repo_prefix", "")

    if not args.skip_clone:
        submissions = Path("repos")
        submissions.mkdir(exist_ok=True)
        for url in urls:
            name = repo_name_from_url(url)
            local = submissions / name
            if local.exists():
                print(f"  pulling {name}...")
                subprocess.run(["git", "-C", str(local), "pull", "--quiet"])
            else:
                print(f"  cloning {name}...")
                subprocess.run(["git", "clone", "--quiet", url, str(local)])

    total_points = config["assignment"].get("total_points", 100)

    rows = []
    for url in urls:
        repo_name = repo_name_from_url(url)
        local = find_local_path(repo_name, args.repos_dir)
        if not local:
            print(f"  WARNING: no local clone for {repo_name}", file=sys.stderr)
            continue

        # Extract github username by stripping the assignment prefix
        github_id = repo_name.removeprefix(prefix) if prefix else repo_name
        display_name = roster.get(github_id, github_id)

        score, reasoning, stats = score_repo(local, config)
        bucket = triage(score, pass_t, flag_t)

        rows.append({
            "name":       display_name,
            "repo_url":   url,
            "triage":     bucket,
            "score":      score,
            "grade":      f"0/{total_points}" if score == 0 and "no commits found" in reasoning else f"/{total_points}",
            "stats":      stats,
            "reasoning":  reasoning,
        })
        print(f"  {bucket:6s}  {score:3d}  {display_name}")

    order = {"FLAG": 0, "REVIEW": 1, "PASS": 2}
    rows.sort(key=lambda r: (order[r["triage"]], -r["score"]))

    # Machine-readable CSV (stats dict excluded — raw values are in reasoning)
    csv_fields = ["name", "repo_url", "triage", "score", "grade", "reasoning"]
    with open(args.output, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=csv_fields, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)

    # Visual spreadsheet
    xlsx_path = Path(args.output).with_suffix(".xlsx")
    write_xlsx(rows, xlsx_path, config)

    # Markdown report
    md_path = Path(args.output).with_suffix(".md")
    write_md(rows, md_path, config)

    counts = {k: sum(1 for r in rows if r["triage"] == k) for k in ["FLAG", "REVIEW", "PASS"]}
    print(f"\nWrote {len(rows)} rows → {args.output}")
    print(f"       {len(rows)} rows → {xlsx_path}")
    print(f"       {len(rows)} rows → {md_path}")
    print(f"  FLAG={counts['FLAG']}  REVIEW={counts['REVIEW']}  PASS={counts['PASS']}")


if __name__ == "__main__":
    main()
